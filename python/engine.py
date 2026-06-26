"""
GSTR 2B Reconciliation — Core Engine
=====================================
All Stage 1 / Stage 2 / Stage 3 logic in one place. This is the single
source of truth for the desktop app: the UI sends raw file bytes here,
this module does every transformation and writes the output .xlsx files.

Nothing in this module reads the filesystem for *input* — inputs arrive as
bytes. Outputs are written to a caller-supplied folder.
"""

import io
import re
from datetime import date, timedelta

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from rapidfuzz import fuzz

# ── Formatting constants (measured from the reference Latest files) ───────────
HEADER_FILL_HEX  = "203764"
HEADER_FONT_COL  = "FFFFFF"
HEADER_FONT_NAME = "Arial"
HEADER_FONT_SIZE = 9
HEADER_ROW_HT    = 24

DATA_FONT_NAME_2B    = "Calibri"
DATA_FONT_SIZE_2B    = 11
DATA_FONT_NAME_TALLY = "Arial"
DATA_FONT_SIZE_TALLY = 9

TALLY_COL_FILL_HEX = "E2EEDA"
ACCOUNTING_FMT = '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ '
DATE_FMT_TALLY = "d-mmm-yy"
DATE_FMT_2B    = "DD/MM/YYYY"

AMT_TOL         = 1.0
FUZZY_THRESHOLD = 90
_EXCEL_EPOCH    = date(1899, 12, 30)

CGST_SGST_COL = "_cgst_sgst"
IGST_COL      = "_igst"

# ── Column specs ──────────────────────────────────────────────────────────────
# 2B: (raw label, output label, halign, number format)
_2B_COLS = [
    ("GSTR-1/IFF/GSTR-5 Period",     "GSTR-1/IFF/GSTR-5 Period",    "left",   "General"),
    ("GSTIN of supplier",             "GSTIN of supplier",            "left",   "@"),
    ("Trade/Legal name",              "Trade/Legal name",             "left",   "@"),
    ("Invoice number",                "Invoice number",               "left",   "@"),
    ("Invoice Date",                  "Invoice Date",                 "left",   DATE_FMT_2B),
    ("Invoice Value(\u20b9)",         "Invoice Value(\u20b9)",        "right",  "0.00"),
    ("Central Tax(\u20b9)",           "Central Tax(\u20b9)",          "right",  "0.00"),
    ("State/UT Tax(\u20b9)",          "State/UT Tax(\u20b9)",         "right",  "0.00"),
    ("Integrated Tax(\u20b9)",        "Integrated Tax(\u20b9)",       "right",  "0.00"),
    ("Cess(\u20b9)",                  "Cess(\u20b9)",                 "right",  "0.00"),
    ("Place of supply",               "Place of supply",              "left",   "General"),
    ("Supply Attract Reverse Charge", "Supply Attract Reverse Charge","center", "General"),
    ("Taxable Value (\u20b9)",        "Taxable Value (\u20b9)",       "right",  "0.00"),
    ("ITC Availability",              "ITC Availability",             "center", "General"),
]
_2B_COL_WIDTHS = {
    "A": 20.0, "B": 13.0, "C": 63.42578125, "D": 21.7109375,
    "E": 15.85546875, "F": 20.0, "G": 13.0, "H": 13.0,
    "I": 23.28515625, "J": 20.0, "K": 13.0, "L": 13.0, "M": 13.0, "N": 13.0,
}
_2B_OUT = [c[1] for c in _2B_COLS]

# Tally: (raw label, output label, halign, number format)
_TALLY_COLS = [
    ("Date",                  "Tally Entry Date",   "right",  DATE_FMT_TALLY),
    ("GSTIN/UIN",             "GSTIN of supplier",  None,     "@"),
    ("Particulars",           "Trade/Legal name",   None,     "@"),
    ("Supplier Invoice No.",  "Invoice number",     None,     "@"),
    ("Supplier Invoice Date", "Invoice Date",       "right",  DATE_FMT_TALLY),
    ("Gross Total",           "Invoice Value(\u20b9)",   "right",  ACCOUNTING_FMT),
    ("Input CGST",            "Central Tax(\u20b9)",     "right",  ACCOUNTING_FMT),
    ("Input SGST",            "State/UT Tax(\u20b9)",    "right",  ACCOUNTING_FMT),
    ("Input IGST",            "Integrated Tax(\u20b9)",  "right",  ACCOUNTING_FMT),
    ("Division",              "Division",           None,     "@"),
]
_TALLY_SUBHEADER = [
    "Date", "GSTIN/UIN", "Particulars", "Supplier Invoice No.",
    "Supplier Invoice Date", "Gross Total", "Input CGST", "Input SGST",
    "Input IGST", "Division",
]
_TALLY_COL_WIDTHS = {
    "A": 14.85546875, "B": 27.7109375, "C": 41.7109375, "D": 20.42578125,
    "E": 20.140625, "F": 21.0, "G": 11.7109375, "H": 13.0, "I": 14.5703125, "J": 13.0,
}
_TALLY_FILLED_COLS = {2, 3, 4, 5, 6, 7, 8, 9}
_TALLY_OUT = [c[1] for c in _TALLY_COLS]


# ══════════════════════════════════════════════════════════════════════════════
# Small helpers
# ══════════════════════════════════════════════════════════════════════════════
def _norm_colname(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val.date()
    if hasattr(val, "date") and callable(val.date):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if s.isdigit():
        return _EXCEL_EPOCH + timedelta(days=int(s))
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return pd.to_datetime(s, format=fmt).date()
        except ValueError:
            continue
    return None


def _strip_x000d(val):
    if val is None:
        return None
    s = str(val).replace("_x000D_", "").replace("\r", "").strip()
    return s if s and s.lower() != "nan" else None


def _is_blank_gstin(series):
    return series.isna() | series.astype(str).str.strip().isin(["", "nan", "NAN", "None"])


def _parse_2b_date(val):
    if pd.isna(val) or str(val).strip() in ("", "nan"):
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return pd.to_datetime(str(val).strip(), format=fmt).date()
        except ValueError:
            continue
    return val


# ══════════════════════════════════════════════════════════════════════════════
# Output formatting
# ══════════════════════════════════════════════════════════════════════════════
def _write_header(ws, headers, row=1):
    fill  = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    font  = Font(name=HEADER_FONT_NAME, size=HEADER_FONT_SIZE, bold=True, color=HEADER_FONT_COL)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, label in enumerate(headers, 1):
        c = ws.cell(row, ci, value=label)
        c.font, c.fill, c.alignment = font, fill, align
    ws.row_dimensions[row].height = HEADER_ROW_HT


def _write_2b_sheet(ws, df, extra=None):
    """extra: list of (col_name, halign, num_fmt) appended after the 14 cols."""
    spec = list(_2B_COLS)
    if extra:
        spec += [(c, c, h, f) for c, h, f in extra]
    headers   = [s[1] for s in spec]
    col_specs = {s[1]: (s[2], s[3]) for s in spec}

    _write_header(ws, headers)
    for letter, w in _2B_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    font = Font(name=DATA_FONT_NAME_2B, size=DATA_FONT_SIZE_2B)
    df = df.reset_index(drop=True)
    for ri, row_data in df.iterrows():
        xl = ri + 2
        for ci, name in enumerate(headers, 1):
            cell = ws.cell(xl, ci)
            val  = row_data.get(name)
            halign, fmt = col_specs[name]
            if fmt == DATE_FMT_2B:
                cell.value = val if isinstance(val, date) else _to_date(val)
                cell.number_format = fmt
            elif fmt == "0.00":
                try:
                    cell.value = float(val) if (pd.notna(val) and str(val).strip() not in ("", "nan")) else None
                except (ValueError, TypeError):
                    cell.value = None
                cell.number_format = fmt
            else:
                s = str(val).strip() if pd.notna(val) else None
                cell.value = None if s in (None, "nan", "") else s
                cell.number_format = fmt
            cell.font = font
            cell.alignment = Alignment(horizontal=halign) if halign else Alignment()


def _write_tally_sheet(ws, df, extra=None, grand_total=False):
    spec = list(_TALLY_COLS)
    subs = list(_TALLY_SUBHEADER)
    if extra:
        for c, h, f in extra:
            spec.append((c, c, h, f))
            subs.append(c)
    headers   = [s[1] for s in spec]
    col_specs = {s[1]: (s[2], s[3]) for s in spec}

    col_fill  = PatternFill("solid", fgColor=TALLY_COL_FILL_HEX)
    data_font = Font(name=DATA_FONT_NAME_TALLY, size=DATA_FONT_SIZE_TALLY, bold=True)
    sub_font  = Font(name=DATA_FONT_NAME_TALLY, size=DATA_FONT_SIZE_TALLY, bold=False)

    _write_header(ws, headers, row=1)
    for ci, text in enumerate(subs, 1):
        cell = ws.cell(2, ci, value=text)
        cell.font = sub_font
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = ACCOUNTING_FMT if ci in {6, 7, 8, 9} else "@"
        if ci in _TALLY_FILLED_COLS:
            cell.fill = col_fill

    for letter, w in _TALLY_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    df = df.reset_index(drop=True)
    for ri, row_data in df.iterrows():
        xl = ri + 3
        for ci, name in enumerate(headers, 1):
            cell = ws.cell(xl, ci)
            val  = row_data.get(name)
            halign, fmt = col_specs[name]
            if fmt == DATE_FMT_TALLY:
                cell.value = val if isinstance(val, date) else _to_date(val)
                cell.number_format = fmt
            elif fmt == ACCOUNTING_FMT:
                try:
                    cell.value = float(str(val).replace(",", "")) if (pd.notna(val) and str(val).strip() not in ("", "nan")) else None
                except (ValueError, TypeError):
                    cell.value = None
                cell.number_format = fmt
            else:
                s = str(val).strip() if pd.notna(val) else None
                cell.value = None if s in (None, "nan", "") else s
                cell.number_format = fmt
            cell.font = data_font
            cell.alignment = Alignment(horizontal=halign if halign else None, vertical="top")
            if ci in _TALLY_FILLED_COLS:
                cell.fill = col_fill

    if grand_total and len(df) > 0:
        gt_fill = PatternFill("solid", fgColor=HEADER_FILL_HEX)
        gt_font = Font(name=DATA_FONT_NAME_TALLY, size=DATA_FONT_SIZE_TALLY, bold=True, color=HEADER_FONT_COL)
        money = {"Invoice Value(\u20b9)", "Central Tax(\u20b9)", "State/UT Tax(\u20b9)", "Integrated Tax(\u20b9)"}
        gt_row = ws.max_row + 1
        for ci, name in enumerate(headers, 1):
            cell = ws.cell(gt_row, ci)
            if name == "Trade/Legal name":
                cell.value = "Grand Total"
                cell.number_format = "@"
            elif name in money:
                tot = pd.to_numeric(df[name], errors="coerce").fillna(0).sum()
                cell.value = round(float(tot), 2)
                cell.number_format = ACCOUNTING_FMT
            else:
                cell.value = None
            cell.font = gt_font
            cell.fill = gt_fill
            cell.alignment = Alignment(horizontal="right", vertical="top")


def _wb_bytes(build_fn):
    """Run build_fn(ws) on a fresh workbook and return xlsx bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    build_fn(ws)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 1 — Clean raw bytes -> cleaned DataFrames
# ══════════════════════════════════════════════════════════════════════════════
def clean_2b(raw_bytes):
    raw_all = pd.read_excel(io.BytesIO(raw_bytes), header=None, dtype=str)
    header_idx = None
    for i, row in raw_all.iterrows():
        if row.astype(str).str.contains("GSTIN of supplier", case=False, na=False).any():
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find the 'GSTIN of supplier' header row in the 2B file.")

    row_a = raw_all.iloc[header_idx]
    row_b = raw_all.iloc[header_idx + 1]
    combined = []
    for a, b in zip(row_a, row_b):
        a_s = str(a).strip() if pd.notna(a) and str(a).strip() not in ("nan", "") else ""
        b_s = str(b).strip() if pd.notna(b) and str(b).strip() not in ("nan", "") else ""
        combined.append(b_s if b_s else a_s)

    df = pd.read_excel(io.BytesIO(raw_bytes), header=None, skiprows=header_idx + 2, dtype=str)
    df.columns = combined[: df.shape[1]]
    rows_read = len(df)

    raw_names = [c[0] for c in _2B_COLS]
    out_names = [c[1] for c in _2B_COLS]
    col_map = {}
    for rn in raw_names:
        for col in df.columns:
            if _norm_colname(col).lower() == _norm_colname(rn).lower():
                col_map[rn] = col
                break
    df = df[[col_map[n] for n in raw_names if n in col_map]].copy()
    df.columns = [out_names[raw_names.index(n)] for n in raw_names if n in col_map]

    if "Invoice Date" in df.columns:
        df["Invoice Date"] = df["Invoice Date"].apply(_parse_2b_date)

    before = len(df)
    df = df[df["GSTIN of supplier"].notna() &
            (~df["GSTIN of supplier"].astype(str).str.strip().isin(["", "nan"]))]
    dropped = before - len(df)

    df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
    df = _add_tax_sort_cols(df)
    df = (df.sort_values(["GSTIN of supplier", "Invoice Date", CGST_SGST_COL, IGST_COL],
                         ascending=[True, True, False, False], kind="stable")
            .drop(columns=[CGST_SGST_COL, IGST_COL]).reset_index(drop=True))

    return df, {"read": rows_read, "dropped": dropped, "written": len(df)}


def clean_tally(raw_bytes):
    df = pd.read_excel(io.BytesIO(raw_bytes), dtype=str)
    rows_read = len(df)

    raw_names = [c[0] for c in _TALLY_COLS]
    out_names = [c[1] for c in _TALLY_COLS]
    col_map = {}
    for rn in raw_names:
        for col in df.columns:
            if _norm_colname(col).lower() == _norm_colname(rn).lower():
                col_map[rn] = col
                break
    df = df[[col_map[n] for n in raw_names if n in col_map]].copy()
    df.columns = [out_names[raw_names.index(n)] for n in raw_names if n in col_map]

    gt_mask = df["Trade/Legal name"].astype(str).str.strip().str.lower() == "grand total"
    df_gt = df[gt_mask].copy()
    df = df[~gt_mask].copy()

    for dc in ("Tally Entry Date", "Invoice Date"):
        if dc in df.columns:
            df[dc] = df[dc].apply(_to_date)
    for tc in ("Trade/Legal name", "Invoice number", "GSTIN of supplier", "Division"):
        if tc in df.columns:
            df[tc] = df[tc].apply(_strip_x000d)

    df["_blank"] = _is_blank_gstin(df["GSTIN of supplier"])
    df = _add_tax_sort_cols(df)
    df = (df.sort_values(["_blank", "GSTIN of supplier", "Invoice Date", CGST_SGST_COL, IGST_COL],
                         ascending=[False, True, True, False, False], kind="stable")
            .drop(columns=["_blank", CGST_SGST_COL, IGST_COL]).reset_index(drop=True))

    blanks = int(_is_blank_gstin(df["GSTIN of supplier"]).sum())
    return df, df_gt, {"read": rows_read, "dropped": int(gt_mask.sum()),
                       "written": len(df), "blanks": blanks}


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 2 — Merge with carry forward + dedup
# ══════════════════════════════════════════════════════════════════════════════
def _dedup_key(df):
    gstin = df["GSTIN of supplier"].astype(str).str.strip().str.upper()
    inv   = df["Invoice number"].astype(str).str.strip().str.upper()
    blank = _is_blank_gstin(df["GSTIN of supplier"])
    name_col = "Trade/Legal name" if "Trade/Legal name" in df.columns else None
    date_col = "Invoice Date" if "Invoice Date" in df.columns else None
    if name_col and date_col:
        name = df[name_col].astype(str).str.strip().str.upper()
        dts  = df[date_col].astype(str).str.strip()
        blank_key = name + "|" + inv + "|" + dts
    else:
        blank_key = inv
    real_key = gstin + "|" + inv
    key = pd.Series(index=df.index, dtype=str)
    key[~blank] = real_key[~blank]
    key[blank]  = blank_key[blank]
    return key


def _coerce_2b_loaded(df):
    df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
    df["Invoice Value(\u20b9)"] = pd.to_numeric(df["Invoice Value(\u20b9)"], errors="coerce")
    for c in ("Central Tax(\u20b9)", "State/UT Tax(\u20b9)", "Integrated Tax(\u20b9)", "Cess(\u20b9)", "Taxable Value (\u20b9)"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df


def _coerce_tally_loaded(df):
    for dc in ("Tally Entry Date", "Invoice Date"):
        if dc in df.columns:
            df[dc] = df[dc].apply(_to_date)
    df["Invoice Value(\u20b9)"] = pd.to_numeric(df["Invoice Value(\u20b9)"], errors="coerce")
    for c in ("Central Tax(\u20b9)", "State/UT Tax(\u20b9)", "Integrated Tax(\u20b9)"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df


def _load_cf_2b(cf_bytes):
    df = pd.read_excel(io.BytesIO(cf_bytes), dtype=str)
    df = df[[c for c in df.columns if c != "Remark"]]
    keep = [c for c in _2B_OUT if c in df.columns]
    df = df[keep].copy()
    return _coerce_2b_loaded(df)


def _load_cf_tally(cf_bytes):
    raw = pd.read_excel(io.BytesIO(cf_bytes), header=None, dtype=str, nrows=2)
    row0 = [str(v).strip() for v in raw.iloc[0]]
    two_header = "Tally Entry Date" in row0
    rename = {
        "Date": "Tally Entry Date", "GSTIN/UIN": "GSTIN of supplier",
        "Particulars": "Trade/Legal name", "Supplier Invoice No.": "Invoice number",
        "Supplier Invoice Date": "Invoice Date", "Gross Total": "Invoice Value(\u20b9)",
        "Input CGST": "Central Tax(\u20b9)", "Input SGST": "State/UT Tax(\u20b9)",
        "Input IGST": "Integrated Tax(\u20b9)",
    }
    if two_header:
        df = pd.read_excel(io.BytesIO(cf_bytes), skiprows=1, dtype=str)
    else:
        df = pd.read_excel(io.BytesIO(cf_bytes), dtype=str)
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "Trade/Legal name" in df.columns:
        df = df[df["Trade/Legal name"].astype(str).str.strip().str.lower() != "grand total"].copy()
    df = df[[c for c in df.columns if c != "Remark"]]
    keep = [c for c in _TALLY_OUT if c in df.columns]
    df = df[keep].copy()
    return _coerce_tally_loaded(df)


def merge_2b(latest_df, cf_bytes):
    if not cf_bytes:
        return latest_df.copy(), {"merged": False, "rows": len(latest_df)}
    cf = _load_cf_2b(cf_bytes)
    combined = pd.concat([cf, latest_df], ignore_index=True)
    key = _dedup_key(combined)
    combined = combined[~key.duplicated(keep="last")].copy()
    combined = _add_tax_sort_cols(combined)
    combined = (combined.sort_values(["GSTIN of supplier", "Invoice Date", CGST_SGST_COL, IGST_COL],
                                     ascending=[True, True, False, False], kind="stable")
                .drop(columns=[CGST_SGST_COL, IGST_COL]).reset_index(drop=True))
    return combined, {"merged": True, "cf_rows": len(cf), "latest_rows": len(latest_df), "rows": len(combined)}


def merge_tally(latest_df, cf_bytes):
    if not cf_bytes:
        return latest_df.copy(), {"merged": False, "rows": len(latest_df)}
    cf = _load_cf_tally(cf_bytes)
    combined = pd.concat([cf, latest_df], ignore_index=True)
    key = _dedup_key(combined)
    combined = combined[~key.duplicated(keep="last")].copy()
    combined["_blank"] = _is_blank_gstin(combined["GSTIN of supplier"])
    combined = _add_tax_sort_cols(combined)
    combined = (combined.sort_values(["_blank", "GSTIN of supplier", "Invoice Date", CGST_SGST_COL, IGST_COL],
                                     ascending=[False, True, True, False, False], kind="stable")
                .drop(columns=["_blank", CGST_SGST_COL, IGST_COL]).reset_index(drop=True))
    return combined, {"merged": True, "cf_rows": len(cf), "latest_rows": len(latest_df), "rows": len(combined)}


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 3 — Reconciliation
# ══════════════════════════════════════════════════════════════════════════════
def _norm_inv(val):
    s = re.sub(r"[\s\-/]", "", str(val).strip().upper())
    return s.lstrip("0") or "0"


def _num(val):
    if pd.isna(val):
        return 0.0
    s = str(val).replace(",", "").strip()
    if s == "" or s.lower() == "nan":
        return 0.0
    return float(s)


def _within(a, b, tol=AMT_TOL):
    av = _num(a)
    bv = _num(b)
    return abs(av - bv) <= tol


def _tax_amounts(row):
    cgst_sgst = (
        _num(row.get("Central Tax(\u20b9)", 0))
        + _num(row.get("State/UT Tax(\u20b9)", 0))
    )
    igst = _num(row.get("Integrated Tax(\u20b9)", 0))
    return cgst_sgst, igst


def _add_tax_sort_cols(df):
    df[CGST_SGST_COL] = (
        pd.to_numeric(df.get("Central Tax(\u20b9)", 0), errors="coerce").fillna(0)
        + pd.to_numeric(df.get("State/UT Tax(\u20b9)", 0), errors="coerce").fillna(0)
    )
    df[IGST_COL] = pd.to_numeric(df.get("Integrated Tax(\u20b9)", 0), errors="coerce").fillna(0)
    return df


def _dates_match(a, b):
    if pd.isna(a) or pd.isna(b):
        return False
    ad = a.date() if isinstance(a, pd.Timestamp) else a
    bd = b.date() if isinstance(b, pd.Timestamp) else b
    return ad == bd


def _inv_match(b_inv, t_inv, exact_only=False):
    bn = _norm_inv(b_inv)
    tn = _norm_inv(t_inv)
    if bn == tn:
        return True
    if exact_only:
        return False
    if bn.isdigit() and tn.isdigit():
        return False
    return fuzz.token_sort_ratio(bn, tn) >= FUZZY_THRESHOLD


def _amounts_match(b, t):
    b_cgst_sgst, b_igst = _tax_amounts(b)
    t_cgst_sgst, t_igst = _tax_amounts(t)
    return _within(t_cgst_sgst, b_cgst_sgst) and _within(t_igst, b_igst)


def _remark(b, t, where):
    parts = []
    b_cgst_sgst, b_igst = _tax_amounts(b)
    t_cgst_sgst, t_igst = _tax_amounts(t)
    d = abs(b_cgst_sgst - t_cgst_sgst)
    if d > AMT_TOL:
        parts.append(f"CGST+SGST differs by \u20b9{d:.2f}")

    d = abs(b_igst - t_igst)
    if d > AMT_TOL:
        parts.append(f"IGST differs by \u20b9{d:.2f}")
    return f"Near-match in {where}: " + (", ".join(parts) if parts else "minor discrepancy")


def reconcile(df_2b, df_tally, excl_set):
    df_2b = df_2b.copy().reset_index(drop=True)
    df_tally = df_tally.copy().reset_index(drop=True)
    df_2b["_g"] = df_2b["GSTIN of supplier"].astype(str).str.strip().str.upper()
    df_tally["_g"] = df_tally["GSTIN of supplier"].astype(str).str.strip().str.upper()

    excl = excl_set or set()
    e2 = set(df_2b[df_2b["_g"].isin(excl)].index)
    et = set(df_tally[df_tally["_g"].isin(excl)].index)

    tally_by_g = {}
    for idx, row in df_tally.iterrows():
        if idx in et:
            continue
        tally_by_g.setdefault(row["_g"], []).append((idx, row))

    matched2b = {}
    matchedt = set()

    def do_pass(exact_only):
        for bi, b in df_2b.iterrows():
            if bi in e2 or bi in matched2b:
                continue
            for ti, t in tally_by_g.get(b["_g"], []):
                if ti in matchedt:
                    continue
                if (_dates_match(b["Invoice Date"], t["Invoice Date"]) and
                        _inv_match(b["Invoice number"], t["Invoice number"], exact_only) and
                        _amounts_match(b, t)):
                    matched2b[bi] = t["Invoice number"]
                    matchedt.add(ti)
                    break

    do_pass(True)
    do_pass(False)

    remarks2b, remarkst = {}, {}
    for bi, b in df_2b.iterrows():
        if bi in e2 or bi in matched2b:
            continue
        best = None
        for ti, t in tally_by_g.get(b["_g"], []):
            if not _dates_match(b["Invoice Date"], t["Invoice Date"]):
                continue
            ex = _inv_match(b["Invoice number"], t["Invoice number"], True)
            fz = (not ex) and _inv_match(b["Invoice number"], t["Invoice number"], False)
            if ex:
                best = t
                break
            if fz and best is None:
                best = t
        if best is not None:
            remarks2b[bi] = _remark(b, best, "Tally")

    b2b_by_g = {}
    for idx, row in df_2b.iterrows():
        if idx in e2:
            continue
        b2b_by_g.setdefault(row["_g"], []).append((idx, row))
    for ti, t in df_tally.iterrows():
        if ti in et or ti in matchedt:
            continue
        best = None
        for bi, b in b2b_by_g.get(t["_g"], []):
            if not _dates_match(t["Invoice Date"], b["Invoice Date"]):
                continue
            ex = _inv_match(t["Invoice number"], b["Invoice number"], True)
            fz = (not ex) and _inv_match(t["Invoice number"], b["Invoice number"], False)
            if ex:
                best = b
                break
            if fz and best is None:
                best = b
        if best is not None:
            remarkst[ti] = _remark(best, t, "2B")

    clean_2b = [c for c in df_2b.columns if not c.startswith("_")]
    clean_t = [c for c in df_tally.columns if not c.startswith("_")]

    df_recon = df_2b.loc[sorted(matched2b), clean_2b].copy().reset_index(drop=True)
    df_recon["Matched Tally Invoice No."] = [matched2b[i] for i in sorted(matched2b)]

    df_e2 = df_2b.loc[sorted(e2), clean_2b].copy()
    df_e2["Source"] = "2B"
    df_et = df_tally.loc[sorted(et), clean_t].copy()
    df_et["Source"] = "Tally"
    tally_only = ["Tally Entry Date", "Division"]
    for c in tally_only:
        if c not in df_e2.columns:
            df_e2[c] = None
    for c in clean_2b:
        if c not in df_et.columns:
            df_et[c] = None
    ntbc_cols = clean_2b + tally_only + ["Source"]
    df_ntbc = pd.concat([df_e2.reindex(columns=ntbc_cols), df_et.reindex(columns=ntbc_cols)], ignore_index=True)

    um2 = [i for i in df_2b.index if i not in e2 and i not in matched2b]
    df_cf2b = df_2b.loc[um2, clean_2b].copy().reset_index(drop=True)
    df_cf2b["Remark"] = [remarks2b.get(i, "") for i in um2]

    umt = [i for i in df_tally.index if i not in et and i not in matchedt]
    df_cft = df_tally.loc[umt, clean_t].copy().reset_index(drop=True)
    df_cft["Remark"] = [remarkst.get(i, "") for i in umt]

    stats = {
        "total2B": len(df_2b), "totalTally": len(df_tally),
        "reconciled": len(df_recon), "ntbc": len(df_ntbc),
        "cf2b": len(df_cf2b), "cfTally": len(df_cft),
        "nearMiss2B": len(remarks2b), "nearMissTally": len(remarkst),
    }
    return df_recon, df_ntbc, df_cf2b, df_cft, stats


def load_exclusion(excl_bytes):
    if not excl_bytes:
        return set()
    df = pd.read_excel(io.BytesIO(excl_bytes), header=None, dtype=str)
    gstins = set(df.iloc[:, 0].dropna().astype(str).str.strip().str.upper()) - {"", "NAN", "GSTIN"}
    return gstins


# ══════════════════════════════════════════════════════════════════════════════
# Output workbook builders -> bytes
# ══════════════════════════════════════════════════════════════════════════════
def build_summary_bytes(df_recon, df_ntbc):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Reconciled Entries"
    _write_2b_sheet(ws1, df_recon, extra=[("Matched Tally Invoice No.", "left", "@")])
    ws2 = wb.create_sheet("Not to be Claimed")
    _write_2b_sheet(ws2, df_ntbc, extra=[
        ("Tally Entry Date", "right", DATE_FMT_2B),
        ("Division", "left", "@"),
        ("Source", "center", "@"),
    ])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_2b_cf_bytes(df):
    return _wb_bytes(lambda ws: _write_2b_sheet(ws, df, extra=[("Remark", "left", "@")]))


def build_tally_cf_bytes(df):
    return _wb_bytes(lambda ws: _write_tally_sheet(ws, df, extra=[("Remark", "left", "@")], grand_total=True))


def build_2b_latest_bytes(df):
    return _wb_bytes(lambda ws: _write_2b_sheet(ws, df))


def build_tally_latest_bytes(df, df_gt=None):
    def build(ws):
        _write_tally_sheet(ws, df, grand_total=False)
        if df_gt is not None and len(df_gt) > 0:
            _append_grand_total(ws, df_gt)
    return _wb_bytes(build)


def _append_grand_total(ws, df_gt):
    gt_fill = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    gt_font = Font(name=DATA_FONT_NAME_TALLY, size=DATA_FONT_SIZE_TALLY, bold=True, color=HEADER_FONT_COL)
    col_specs = {c[1]: (c[2], c[3]) for c in _TALLY_COLS}
    for _, gt in df_gt.iterrows():
        xl = ws.max_row + 1
        for ci, name in enumerate(_TALLY_OUT, 1):
            cell = ws.cell(xl, ci)
            _, fmt = col_specs[name]
            if name == "Trade/Legal name":
                cell.value = "Grand Total"
                cell.number_format = "@"
            elif fmt == ACCOUNTING_FMT:
                try:
                    v = gt.get(name)
                    cell.value = float(str(v).replace(",", "")) if (pd.notna(v) and str(v).strip() not in ("", "nan")) else None
                except (ValueError, TypeError):
                    cell.value = None
                cell.number_format = fmt
            else:
                cell.value = None
                cell.number_format = fmt
            cell.font = gt_font
            cell.fill = gt_fill
            cell.alignment = Alignment(horizontal="right" if ci in _TALLY_FILLED_COLS else None, vertical="top")


# ══════════════════════════════════════════════════════════════════════════════
# Preview helper — DataFrame -> list[dict] of display strings (for the UI)
# ══════════════════════════════════════════════════════════════════════════════
def preview_records(df, limit=8):
    out = []
    for _, row in df.head(limit).iterrows():
        rec = {}
        for col, val in row.items():
            if isinstance(val, (pd.Timestamp, date)):
                d = val.date() if isinstance(val, pd.Timestamp) else val
                rec[col] = f"{d.day:02d}/{d.month:02d}/{d.year}"
            elif pd.isna(val):
                rec[col] = ""
            elif isinstance(val, float):
                rec[col] = f"{val:.2f}"
            else:
                rec[col] = str(val)
        out.append(rec)
    return out
